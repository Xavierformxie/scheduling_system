from PyQt6.QtWidgets import QFileDialog, QMessageBox, QTableWidgetItem, QComboBox, QLineEdit
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


class ExcelHandler:
    @staticmethod
    def import_staff(staff_table):
        file_name, _ = QFileDialog.getOpenFileName(None,"选择Excel文件","","Excel Files(*xlsx *.xls)")
        if file_name:
            try:
                df = pd.read_excel(file_name)
                staff_table.setRowCount(0)

                #获取显示的列名和对应的键
                display_to_key = {col['display']: col['key'] for col in staff_table.config['columns']}

                # 确保 DataFrame 的列名与 staff_table 的显示列名一致
                df = df.reindex(columns=list(display_to_key.keys()))

                #将NaN 值替换为空字符串
                df = df.fillna('')

                staff_table.staff_data = []
                for _, row in df.iterrows():
                    staff = {}
                    for display_column, key in display_to_key.items():
                        column_config = next(col for col in staff_table.config['columns'] if col['key'] == key)
                        if column_config.get('type') == 'list':
                            if pd.notna(row[display_column]) and not isinstance(row[display_column], list):
                                items = staff_table.split_list_items(str(row[display_column]), column_config)
                                staff[key] = "、".join(items)  # 将列表重新拼接为字符串
                            else:
                                staff[key] = ""  # 将空值设置为空字符串
                        else:
                            staff[key] = str(row[display_column])
                    # 确保空列表不会显示为 "[]"
                    for key, value in staff.items():
                        if isinstance(value, list) and not value:
                            staff[key] = ""
                    staff_table.staff_data.append(staff)

                staff_table.setup_table()
                staff_table.populate_table()
                staff_table.staff_data_changed.emit()

                QMessageBox.information(None,"导入成功","员工信息已成功导入")

            except Exception as e:
                QMessageBox.critical(None,"导入失败",f"导入过程中出现错误:{str(e)}")

    @staticmethod
    def export_staff(staff_table):
        file_name, _ = QFileDialog.getSaveFileName(None,"保存Excel文件","","Excel Files(*.xlsx)")
        if file_name:
            try:
                wb = Workbook()
                ws = wb.active

                #写入表头
                for col, column_config in enumerate(staff_table.config['columns'], start=1):
                    ws.cell(row=1, column=col, value=column_config['display'])

                #写入数据并设置下拉列表
                for row in range(staff_table.rowCount()):
                    for col, column_config in enumerate(staff_table.config['columns'], start=1):
                        cell = ws.cell(row=row+2, column=col)
                        widget = staff_table.cellWidget(row,col-1)

                        if column_config.get('type')== 'select':
                            if isinstance(widget, QComboBox):
                                value = widget.currentText()
                            else:
                                value = staff_table.item(row, col-1).text()
                            cell.value = value
                            options = column_config.get('options',[])
                            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"',allow_blank=True)
                            ws.add_data_validation(dv)
                            dv.add(cell)
                        elif column_config.get('type')== 'list':
                            value = widget.text() if isinstance(widget, QLineEdit) else staff_table.item(row, col-1).text()
                            cell.value = "" if value == "[]" else value
                        else:
                            value = widget.text() if isinstance(widget, QLineEdit) else staff_table.item(row, col-1).text()
                            cell.value = value

                wb.save(file_name)
                QMessageBox.information(None,"导出成功",f"员工信息已成功导出到 {file_name}")

            except PermissionError:
                QMessageBox.critical(None,"导出失败","无法写入文件。请确保文件没有被其他程序打开，并且您有写入权限。")
            except Exception as e:
                QMessageBox.critical(None,"导入失败",f"导入过程中出现错误:{str(e)}")

    @staticmethod
    def export_result(staff_table):
        pass